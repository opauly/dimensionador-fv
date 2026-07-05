"""Parse ARESEP Cuadro E-8 xlsx → structured T-RE and T-CO tariff data per distributor.

The workbook has one sheet per distributor per year plus 'Vigentes' summary sheets.
This module reads only the '*vig' sheets and extracts:
  T-RE  Tarifa Residencial  (residential, sector='residential')
  T-CO  Comercios y Servicios (commercial, sector='commercial')

T-RE has two ARESEP sub-formats:
  Simple  (CNFL, JASEC, CG, CS, CAR): a.=fixed min charge, b./c./d.=per-kWh tiers
  Paired  (ICE, ESPH, CL):            a/c/e/g=block minimums, b/d/f/h=per-kWh rates

T-CO has two sub-formats:
  5-row (most distributors): a=energy rate ≤3000 kWh, b=min charge, c=energy rate >3000 kWh,
                              d=demand min charge, e=demand rate per kW
  3-row (ICE):               a=energy rate ≤3000 kWh, b=energy rate per kWh, c=demand rate per kW
"""
from __future__ import annotations
import re
import statistics
from typing import IO

import openpyxl

SHEET_TO_ABBREV: dict[str, str] = {
    "ICE vig":   "ICE",
    "CNFL vig":  "CNFL",
    "JASEC vig": "JASEC",
    "ESPH vig":  "ESPH",
    "CL vig":    "COOPELESCA",
    "CG vig":    "COOPEGUANACASTE",
    "CS vig":    "COOPESANTOS",
    "CAR vig":   "COOPEALFARORUIZ",
}

_LETTERS = list("abcdefghijklmnop")


def _row_letter(label: str) -> str | None:
    m = re.match(r"^([a-z])\.\s", label.strip().lower())
    return m.group(1) if m else None


def _avg_row_value(row: tuple, start: int = 1, end: int = 13) -> float | None:
    vals = [v for v in row[start:end] if isinstance(v, (int, float))]
    return statistics.mean(vals) if vals else None


def _parse_kwh_range(label: str) -> tuple[int, int | None]:
    """Parse 'from' and 'to' kWh from a block label string."""
    s = label.lower()
    m = re.search(r"mayor\s+a\s+([\d\s]+)\s*k[wh]+", s)
    if m:
        return int(m.group(1).replace(" ", "")) + 1, None
    m = re.search(r"([\d]+)\s*-\s*([\d]+)\s*k[wh]+", s)
    if m:
        return int(m.group(1)), int(m.group(2))
    m = re.search(r"([\d]+)\s*k[wh]+", s)
    if m:
        return 0, int(m.group(1))
    return 0, None


def _parse_kw_threshold(label: str) -> int:
    """Extract kW threshold from a demand block label like 'Bloque 0-8 kW'."""
    s = label.lower()
    m = re.search(r"0\s*-\s*([\d]+)\s*k[w](?!h)", s)
    if m:
        return int(m.group(1))
    m = re.search(r"([\d]+)\s*k[w](?!h)", s)
    if m:
        return int(m.group(1))
    return 0


def _is_paired_format(rows: list[tuple[str, float]]) -> bool:
    """Return True if two consecutive rows share the same kWh range (T-RE paired format)."""
    for i in range(len(rows) - 1):
        a_range = re.search(r"[\d]+-[\d]+\s*k[wh]+", rows[i][0].lower())
        b_range = re.search(r"[\d]+-[\d]+\s*k[wh]+", rows[i + 1][0].lower())
        if a_range and b_range and a_range.group() == b_range.group():
            return True
    return False


# ── T-RE parsers ──────────────────────────────────────────────────────────────

def _parse_tre_simple(rows: list[tuple[str, float]]) -> dict:
    """Simple format: row 0 = access charge, rows 1+ = per-kWh energy tiers."""
    access_charge = round(rows[0][1], 2)
    tiers = []
    for label, val in rows[1:]:
        from_kwh, to_kwh = _parse_kwh_range(label)
        tiers.append({
            "from_kwh": from_kwh,
            "to_kwh": to_kwh,
            "rate_crc": round(val, 4),
            "sort_order": len(tiers) + 1,
            "is_fixed": False,
        })
    return {"access_charge_crc": access_charge, "tiers": tiers}


def _parse_tre_paired(rows: list[tuple[str, float]]) -> dict:
    """Paired format: even-letter rows (a,c,e,g) = block minimums, odd (b,d,f,h) = rates."""
    access_charge: float | None = None
    tiers = []
    for label, val in rows:
        letter = _row_letter(label)
        if letter not in _LETTERS:
            continue
        idx = _LETTERS.index(letter)
        if idx % 2 == 0:
            if letter == "a":
                access_charge = round(val, 2)
        else:
            from_kwh, to_kwh = _parse_kwh_range(label)
            tiers.append({
                "from_kwh": from_kwh,
                "to_kwh": to_kwh,
                "rate_crc": round(val, 4),
                "sort_order": len(tiers) + 1,
                "is_fixed": False,
            })
    if tiers:
        tiers[-1]["to_kwh"] = None
    return {"access_charge_crc": access_charge, "tiers": tiers}


def parse_tre_from_sheet(ws) -> dict | None:
    """Extract T-RE tariff from a single vig worksheet. Returns None if not found."""
    rows: list[tuple[str, float]] = []
    in_tre = False

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True):
        label = row[0]
        if label is None:
            continue
        label_s = str(label).strip()
        if label_s == "T-RE":
            in_tre = True
            rows = []
            continue
        if in_tre:
            if _row_letter(label_s) is None:
                break
            val = _avg_row_value(row)
            if val is not None:
                rows.append((label_s, val))

    if not rows:
        return None

    paired = _is_paired_format(rows)
    result = _parse_tre_paired(rows) if paired else _parse_tre_simple(rows)
    result["paired_format"] = paired
    return result


# ── T-CO parsers ──────────────────────────────────────────────────────────────

def parse_tco_from_sheet(ws) -> dict | None:
    """
    Extract T-CO (commercial) tariff from a vig worksheet.

    Returns:
        {
            access_charge_crc: float       — minimum energy bill (b. block 0-3000 kWh)
            demand_rate_crc: float         — CRC per kW above threshold (e. row)
            demand_threshold_kw: int       — kW threshold for demand charge (d. row label)
            tiers: list[dict]              — energy tiers (kWh-based)
        }
    """
    rows: list[tuple[str, float]] = []
    in_tco = False

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True):
        label = row[0]
        if label is None:
            continue
        label_s = str(label).strip()
        if label_s == "T-CO":
            in_tco = True
            rows = []
            continue
        if in_tco:
            if _row_letter(label_s) is None:
                break
            val = _avg_row_value(row)
            if val is not None:
                rows.append((label_s, val))

    if not rows:
        return None

    # ICE simplified format (3 rows): a=rate≤3000, b=rate>3000, c=demand rate
    if len(rows) == 3:
        rate_tier1 = round(rows[0][1], 4)   # a: energy rate ≤3000 kWh
        rate_tier2 = round(rows[1][1], 4)   # b: energy rate per kWh (>3000)
        demand_rate = round(rows[2][1], 2)   # c: demand rate per kW
        return {
            "access_charge_crc": 0.0,
            "demand_rate_crc": demand_rate,
            "demand_threshold_kw": 0,
            "tiers": [
                {"from_kwh": 0,    "to_kwh": 3000, "rate_crc": rate_tier1, "sort_order": 1, "is_fixed": False},
                {"from_kwh": 3001, "to_kwh": None, "rate_crc": rate_tier2, "sort_order": 2, "is_fixed": False},
            ],
        }

    # Standard 5-row format (CNFL, JASEC, ESPH, CL, CG, CS, CAR):
    # a: energy rate ≤3000 kWh  b: min energy charge (0-3000 block)
    # c: energy rate >3000 kWh  d: demand min charge (0-X kW)  e: demand rate per kW
    rate_tier1     = round(rows[0][1], 4)   # a
    access_charge  = round(rows[1][1], 2)   # b — minimum monthly energy charge
    rate_tier2     = round(rows[2][1], 4)   # c
    # d row: demand minimum charge (we ignore the flat floor charge)
    demand_label   = rows[3][0]             # d label, e.g. "d. Bloque 0-8 kW"
    demand_threshold = _parse_kw_threshold(demand_label)
    demand_rate    = round(rows[4][1], 2)   # e — CRC per kW above threshold

    return {
        "access_charge_crc": access_charge,
        "demand_rate_crc": demand_rate,
        "demand_threshold_kw": demand_threshold,
        "tiers": [
            {"from_kwh": 0,    "to_kwh": 3000, "rate_crc": rate_tier1, "sort_order": 1, "is_fixed": False},
            {"from_kwh": 3001, "to_kwh": None, "rate_crc": rate_tier2, "sort_order": 2, "is_fixed": False},
        ],
    }


# ── public API ────────────────────────────────────────────────────────────────

def parse_vigentes(file: IO[bytes] | str) -> dict[str, dict]:
    """
    Parse ARESEP Cuadro E-8 xlsx.

    Returns dict keyed by distributor abbreviation. Each value contains:
        "T-RE": {access_charge_crc, tiers, paired_format}
        "T-CO": {access_charge_crc, demand_rate_crc, demand_threshold_kw, tiers}

    Example:
        result["CNFL"]["T-RE"]["tiers"]   → residential energy tiers
        result["CNFL"]["T-CO"]["demand_rate_crc"]  → commercial demand rate
    """
    wb = openpyxl.load_workbook(file, data_only=True)
    result: dict[str, dict] = {}

    for sheet_name, abbrev in SHEET_TO_ABBREV.items():
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        entry: dict[str, dict] = {}

        tre = parse_tre_from_sheet(ws)
        if tre:
            entry["T-RE"] = tre

        tco = parse_tco_from_sheet(ws)
        if tco:
            entry["T-CO"] = tco

        if entry:
            result[abbrev] = entry

    return result
